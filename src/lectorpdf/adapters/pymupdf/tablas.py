"""Adaptador de `ServicioTablas` sobre PyMuPDF (`find_tables`).

Las celdas vacías llegan como None y se escriben como cadena vacía, y los saltos
de línea dentro de una celda se colapsan: un CSV con saltos crudos dentro de una
celda es un CSV que muchos programas leen mal.
"""

from __future__ import annotations

import csv
import os
from pathlib import Path

import fitz

from lectorpdf.adapters.pymupdf.registro import RegistroDocumentos
from lectorpdf.core.domain.contenido import PalabraTexto
from lectorpdf.core.domain.formularios import RectanguloPt
from lectorpdf.core.domain.herramientas import Progreso
from lectorpdf.core.domain.tablas import (
    EstrategiaTablas,
    FormatoTabla,
    TablaDetectada,
    filas_desde_palabras,
)


class PyMuPDFTablas:
    def __init__(self, registro: RegistroDocumentos) -> None:
        self._registro = registro

    def detectar_tablas(
        self,
        documento_id: str,
        estrategia: EstrategiaTablas = EstrategiaTablas.LINEAS,
        progreso: Progreso | None = None,
    ) -> tuple[TablaDetectada, ...]:
        doc = self._registro.obtener(documento_id)
        detectadas: list[TablaDetectada] = []
        total = doc.page_count
        for i in range(total):
            for j, tabla in enumerate(self._tablas_de(doc[i], estrategia)):
                detectadas.append(
                    TablaDetectada(
                        pagina=i,
                        indice=j,
                        filas=int(tabla.row_count),
                        columnas=int(tabla.col_count),
                    )
                )
            if progreso is not None:
                progreso(i + 1, total)
        return tuple(detectadas)

    def exportar_tablas(
        self,
        documento_id: str,
        directorio: Path,
        formato: FormatoTabla,
        estrategia: EstrategiaTablas = EstrategiaTablas.LINEAS,
        progreso: Progreso | None = None,
    ) -> list[Path]:
        doc = self._registro.obtener(documento_id)
        base = Path(doc.name).stem or "documento"
        directorio.mkdir(parents=True, exist_ok=True)

        contenidos: list[tuple[TablaDetectada, list[list[str]]]] = []
        total = doc.page_count
        for i in range(total):
            for j, tabla in enumerate(self._tablas_de(doc[i], estrategia)):
                filas = [[_celda(c) for c in fila] for fila in tabla.extract()]
                contenidos.append(
                    (
                        TablaDetectada(i, j, len(filas), len(filas[0]) if filas else 0),
                        filas,
                    )
                )
            if progreso is not None:
                progreso(i + 1, total)

        if not contenidos:
            return []
        if formato is FormatoTabla.CSV:
            return _escribir_csv(contenidos, directorio, base)
        return _escribir_xlsx(contenidos, directorio, base)

    def exportar_texto_como_hoja(
        self,
        documento_id: str,
        destino: Path,
        formato: FormatoTabla,
        progreso: Progreso | None = None,
    ) -> Path | None:
        doc = self._registro.obtener(documento_id)
        paginas: list[tuple[int, list[list[str]]]] = []
        total = doc.page_count
        for i in range(total):
            filas = filas_desde_palabras(_palabras_de(doc[i]))
            if filas:
                paginas.append((i, filas))
            if progreso is not None:
                progreso(i + 1, total)

        if not paginas:
            return None
        destino.parent.mkdir(parents=True, exist_ok=True)
        if formato is FormatoTabla.CSV:
            _escribir_csv_seguido(paginas, destino)
        else:
            _escribir_xlsx_por_pagina(paginas, destino)
        return destino

    # -- Interno ------------------------------------------------------------

    @staticmethod
    def _tablas_de(pagina: fitz.Page, estrategia: EstrategiaTablas) -> list[fitz.table.Table]:
        encontradas = pagina.find_tables(
            vertical_strategy=estrategia.value, horizontal_strategy=estrategia.value
        )
        return list(encontradas.tables)


def _palabras_de(pagina: fitz.Page) -> list[PalabraTexto]:
    """Palabras de la página en orden de lectura, como las quiere el dominio.
    Cada tupla de `words` es (x0, y0, x1, y1, texto, bloque, línea, palabra)."""
    return [
        PalabraTexto(RectanguloPt(x0, y0, x1, y1), texto, bloque, linea)
        for x0, y0, x1, y1, texto, bloque, linea, _ in pagina.get_text("words", sort=True)
    ]


def _escribir_csv_seguido(paginas: list[tuple[int, list[list[str]]]], destino: Path) -> None:
    """Un solo CSV con todo el documento, una página detrás de otra. Sin
    columna de página ni separadores: un CSV con adornos es un CSV que luego
    hay que limpiar a mano. En Excel las páginas sí van separadas, en hojas."""
    tmp = destino.with_name(destino.name + ".tmp")
    # utf-8-sig para que Excel abra los acentos bien al hacer doble clic.
    with open(tmp, "w", encoding="utf-8-sig", newline="") as fichero:
        escritor = csv.writer(fichero, delimiter=";")
        for _, filas in paginas:
            escritor.writerows(filas)
    os.replace(tmp, destino)


def _escribir_xlsx_por_pagina(
    paginas: list[tuple[int, list[list[str]]]], destino: Path
) -> None:
    from openpyxl import Workbook

    libro = Workbook()
    libro.remove(libro.active)  # la hoja vacía que trae de fábrica
    for numero, filas in paginas:
        hoja = libro.create_sheet(title=f"Página {numero + 1}")
        for fila in filas:
            hoja.append(fila)
    tmp = destino.with_name(destino.name + ".tmp")
    libro.save(str(tmp))
    os.replace(tmp, destino)


def _celda(valor: object) -> str:
    if valor is None:
        return ""
    return " ".join(str(valor).split())


def _escribir_csv(
    contenidos: list[tuple[TablaDetectada, list[list[str]]]],
    directorio: Path,
    base: str,
) -> list[Path]:
    rutas: list[Path] = []
    for tabla, filas in contenidos:
        salida = directorio / f"{base}_{tabla.nombre}.csv"
        # utf-8-sig para que Excel abra los acentos bien al hacer doble clic.
        with open(salida, "w", encoding="utf-8-sig", newline="") as fichero:
            csv.writer(fichero, delimiter=";").writerows(filas)
        rutas.append(salida)
    return rutas


def _escribir_xlsx(
    contenidos: list[tuple[TablaDetectada, list[list[str]]]],
    directorio: Path,
    base: str,
) -> list[Path]:
    from openpyxl import Workbook

    libro = Workbook()
    libro.remove(libro.active)  # la hoja vacía que trae de fábrica
    for tabla, filas in contenidos:
        hoja = libro.create_sheet(title=tabla.nombre)
        for fila in filas:
            hoja.append(fila)
    salida = directorio / f"{base}_tablas.xlsx"
    libro.save(str(salida))
    return [salida]
