"""Tests de detección y conversión de tablas (PyMuPDF find_tables).

El fixture tiene dos tablas a propósito: una dibujada con líneas y otra que solo
está alineada con espacios, que es como vienen muchas tablas reales.
"""

from __future__ import annotations

import csv
from pathlib import Path

import fitz

from lectorpdf.adapters.pymupdf.document_repository import PyMuPDFDocumentRepository
from lectorpdf.adapters.pymupdf.registro import RegistroDocumentos
from lectorpdf.adapters.pymupdf.tablas import PyMuPDFTablas
from lectorpdf.core.domain.tablas import EstrategiaTablas, FormatoTabla

_FILAS = [
    ["Concepto", "Importe", "Fecha"],
    ["Alquiler", "750,00", "07-09-2026"],
    ["Suministros", "120,50", "05-10-2026"],
]


def _pdf_con_tablas(destino: Path) -> Path:
    """Página 1: tabla con líneas. Página 2: la misma, solo alineada."""
    doc = fitz.open()
    pagina = doc.new_page(width=595, height=842)
    x0, y0, ancho, alto = 60.0, 80.0, 150.0, 24.0
    for f, fila in enumerate(_FILAS):
        for c, celda in enumerate(fila):
            rect = fitz.Rect(
                x0 + c * ancho, y0 + f * alto, x0 + (c + 1) * ancho, y0 + (f + 1) * alto
            )
            pagina.draw_rect(rect, color=(0, 0, 0), width=0.7)
            pagina.insert_text((rect.x0 + 4, rect.y1 - 8), celda, fontsize=10)

    sin_lineas = doc.new_page(width=595, height=842)
    for f, fila in enumerate(_FILAS):
        for c, celda in enumerate(fila):
            sin_lineas.insert_text(
                (x0 + c * ancho, y0 + f * alto), celda, fontsize=10
            )
    doc.save(destino)
    doc.close()
    return destino


def _servicio(ruta: Path) -> tuple[PyMuPDFTablas, str]:
    registro = RegistroDocumentos()
    documento = PyMuPDFDocumentRepository(registro).abrir(ruta)
    return PyMuPDFTablas(registro), documento.id


def test_detecta_la_tabla_con_lineas_y_no_inventa_otras(tmp_path: Path) -> None:
    servicio, doc_id = _servicio(_pdf_con_tablas(tmp_path / "t.pdf"))

    detectadas = servicio.detectar_tablas(doc_id, EstrategiaTablas.LINEAS)

    assert len(detectadas) == 1  # la de la página 2 no tiene líneas: no se ve
    tabla = detectadas[0]
    assert tabla.pagina == 0
    assert tabla.filas == 3 and tabla.columnas == 3
    assert tabla.nombre == "p1_t1"


def test_la_estrategia_de_texto_encuentra_la_tabla_sin_lineas(tmp_path: Path) -> None:
    servicio, doc_id = _servicio(_pdf_con_tablas(tmp_path / "t.pdf"))

    detectadas = servicio.detectar_tablas(doc_id, EstrategiaTablas.TEXTO)

    assert {t.pagina for t in detectadas} == {0, 1}  # también la alineada


def test_csv_con_las_celdas_del_documento(tmp_path: Path) -> None:
    servicio, doc_id = _servicio(_pdf_con_tablas(tmp_path / "t.pdf"))
    salida = tmp_path / "csv"

    rutas = servicio.exportar_tablas(doc_id, salida, FormatoTabla.CSV)

    assert len(rutas) == 1
    assert rutas[0].name == "t_p1_t1.csv"
    with open(rutas[0], encoding="utf-8-sig", newline="") as fichero:
        filas = list(csv.reader(fichero, delimiter=";"))
    assert filas == _FILAS


def test_xlsx_reune_las_tablas_en_un_libro(tmp_path: Path) -> None:
    servicio, doc_id = _servicio(_pdf_con_tablas(tmp_path / "t.pdf"))

    rutas = servicio.exportar_tablas(
        doc_id, tmp_path, FormatoTabla.XLSX, EstrategiaTablas.TEXTO
    )

    assert len(rutas) == 1 and rutas[0].suffix == ".xlsx"

    from openpyxl import load_workbook

    libro = load_workbook(rutas[0])
    try:
        assert len(libro.sheetnames) >= 2  # una hoja por tabla detectada
        assert libro.sheetnames[0] == "p1_t1"
        primera = libro["p1_t1"]
        assert [c.value for c in primera[1]] == _FILAS[0]
    finally:
        libro.close()


def test_documento_sin_tablas_no_escribe_nada(tmp_path: Path) -> None:
    ruta = tmp_path / "vacio.pdf"
    doc = fitz.open()
    doc.new_page().insert_text((72, 100), "Solo un párrafo suelto.", fontsize=11)
    doc.save(ruta)
    doc.close()
    servicio, doc_id = _servicio(ruta)
    salida = tmp_path / "salida"

    rutas = servicio.exportar_tablas(doc_id, salida, FormatoTabla.CSV)

    assert rutas == []
    assert not any(salida.glob("*.csv"))


# -- Volcado del texto entero a hoja de cálculo -----------------------------
#
# Aquí no se detecta nada: se comprueba que lo que sale se parece al documento
# línea a línea, incluida la tabla que la detección por líneas no ve.


def test_csv_del_texto_entero_respeta_lineas_y_columnas(tmp_path: Path) -> None:
    servicio, doc_id = _servicio(_pdf_con_tablas(tmp_path / "t.pdf"))
    destino = tmp_path / "todo.csv"

    escrito = servicio.exportar_texto_como_hoja(doc_id, destino, FormatoTabla.CSV)

    assert escrito == destino
    with open(destino, encoding="utf-8-sig", newline="") as fichero:
        filas = list(csv.reader(fichero, delimiter=";"))
    # Las dos páginas del fixture llevan las mismas tres filas.
    assert filas[:3] == _FILAS
    assert len(filas) == 2 * len(_FILAS)


def test_xlsx_del_texto_entero_pone_una_hoja_por_pagina(tmp_path: Path) -> None:
    servicio, doc_id = _servicio(_pdf_con_tablas(tmp_path / "t.pdf"))
    destino = tmp_path / "todo.xlsx"

    escrito = servicio.exportar_texto_como_hoja(doc_id, destino, FormatoTabla.XLSX)

    assert escrito == destino

    from openpyxl import load_workbook

    libro = load_workbook(destino)
    try:
        assert libro.sheetnames == ["Página 1", "Página 2"]
        hoja = libro["Página 1"]
        assert [c.value for c in hoja[1]] == _FILAS[0]
        assert hoja.max_row == len(_FILAS)
    finally:
        libro.close()


def test_un_parrafo_corriente_no_se_trocea_en_celdas(tmp_path: Path) -> None:
    ruta = tmp_path / "parrafo.pdf"
    doc = fitz.open()
    doc.new_page().insert_text((72, 100), "Solo un párrafo suelto.", fontsize=11)
    doc.save(ruta)
    doc.close()
    servicio, doc_id = _servicio(ruta)
    destino = tmp_path / "p.csv"

    servicio.exportar_texto_como_hoja(doc_id, destino, FormatoTabla.CSV)

    with open(destino, encoding="utf-8-sig", newline="") as fichero:
        filas = list(csv.reader(fichero, delimiter=";"))
    assert filas == [["Solo un párrafo suelto."]]


def test_un_pdf_sin_texto_no_deja_fichero(tmp_path: Path) -> None:
    ruta = tmp_path / "sin_texto.pdf"
    doc = fitz.open()
    doc.new_page()  # una página en blanco: ni una palabra
    doc.save(ruta)
    doc.close()
    servicio, doc_id = _servicio(ruta)
    destino = tmp_path / "vacio.xlsx"

    assert servicio.exportar_texto_como_hoja(doc_id, destino, FormatoTabla.XLSX) is None
    assert not destino.exists()
