"""Criterio de aceptación funcional del volcado a Excel/CSV, sin UI ni red.

Convierte un PDF de fixture entero a XLSX y a CSV y comprueba, celda a celda,
que el resultado se parece al documento: las filas en su orden, las columnas
donde el papel deja un hueco, el párrafo corriente sin trocear, y la tabla
alineada solo con espacios — la que la detección por líneas no ve — también
puesta en columnas. Y que un PDF sin texto no deja fichero.

Uso:
    uv run python scripts/verificar_texto_a_hoja.py
"""

from __future__ import annotations

import csv
import sys
import tempfile
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # consola cp1252 → utf-8

import fitz  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from lectorpdf.adapters.pymupdf.document_repository import (  # noqa: E402
    PyMuPDFDocumentRepository,
)
from lectorpdf.adapters.pymupdf.registro import RegistroDocumentos  # noqa: E402
from lectorpdf.adapters.pymupdf.tablas import PyMuPDFTablas  # noqa: E402
from lectorpdf.core.domain.tablas import FormatoTabla  # noqa: E402
from lectorpdf.core.use_cases.convertir_texto_a_hoja import (  # noqa: E402
    ConvertirTextoAHoja,
)

_resultados: list[tuple[str, bool]] = []

_FILAS = [
    ["Concepto", "Importe", "Fecha"],
    ["Alquiler", "750,00", "07-09-2026"],
    ["Suministros", "120,50", "05-10-2026"],
]
_PARRAFO = "Este párrafo corriente no es una tabla y no debe trocearse."


def _check(nombre: str, condicion: bool) -> None:
    _resultados.append((nombre, bool(condicion)))


def _pdf_de_prueba(destino: Path) -> Path:
    """Página 1: tabla con líneas y un párrafo debajo. Página 2: la misma tabla
    alineada solo con espacios, que es como vienen muchas tablas reales."""
    doc = fitz.open()
    x0, y0, ancho, alto = 60.0, 80.0, 150.0, 24.0

    pagina = doc.new_page(width=595, height=842)
    for f, fila in enumerate(_FILAS):
        for c, celda in enumerate(fila):
            rect = fitz.Rect(
                x0 + c * ancho, y0 + f * alto, x0 + (c + 1) * ancho, y0 + (f + 1) * alto
            )
            pagina.draw_rect(rect, color=(0, 0, 0), width=0.7)
            pagina.insert_text((rect.x0 + 4, rect.y1 - 8), celda, fontsize=10)
    pagina.insert_text((x0, y0 + len(_FILAS) * alto + 40), _PARRAFO, fontsize=11)

    sin_lineas = doc.new_page(width=595, height=842)
    for f, fila in enumerate(_FILAS):
        for c, celda in enumerate(fila):
            sin_lineas.insert_text((x0 + c * ancho, y0 + f * alto), celda, fontsize=10)

    doc.save(destino)
    doc.close()
    return destino


def main() -> int:
    with tempfile.TemporaryDirectory() as carpeta:
        tmp = Path(carpeta)
        registro = RegistroDocumentos()
        repositorio = PyMuPDFDocumentRepository(registro)
        caso = ConvertirTextoAHoja(PyMuPDFTablas(registro))

        documento = repositorio.abrir(_pdf_de_prueba(tmp / "hoja.pdf"))

        # -- Excel ------------------------------------------------------------
        xlsx = caso.ejecutar(documento, tmp / "todo.xlsx", FormatoTabla.XLSX)
        _check("PDF→XLSX: el fichero existe y pesa", xlsx.exists() and xlsx.stat().st_size > 0)

        from openpyxl import load_workbook

        libro = load_workbook(xlsx)
        try:
            _check(
                "XLSX: una hoja por página, nombrada por su número",
                libro.sheetnames == ["Página 1", "Página 2"],
            )
            hoja = libro["Página 1"]
            # iter_rows rellena con None hasta el ancho de la hoja: las filas
            # cortas (un párrafo suelto) llevan huecos que no son celdas.
            leidas = [
                [c.value for c in fila if c.value is not None] for fila in hoja.iter_rows()
            ]
            _check("XLSX: la tabla con líneas, celda a celda", leidas[:3] == _FILAS)
            _check(
                "XLSX: el párrafo corriente queda en una sola celda",
                leidas[3] == [_PARRAFO],
            )
            alineada = [[c.value for c in fila] for fila in libro["Página 2"].iter_rows()]
            _check(
                "XLSX: la tabla sin líneas también sale en columnas",
                alineada == _FILAS,
            )
        finally:
            libro.close()

        # -- CSV --------------------------------------------------------------
        ruta_csv = caso.ejecutar(documento, tmp / "todo.csv", FormatoTabla.CSV)
        with open(ruta_csv, encoding="utf-8-sig", newline="") as fichero:
            filas = list(csv.reader(fichero, delimiter=";"))
        _check("PDF→CSV: la tabla con líneas, celda a celda", filas[:3] == _FILAS)
        _check("CSV: el documento entero, página tras página", filas[4:7] == _FILAS)
        _check(
            "CSV: los acentos se leen bien (utf-8 con BOM para Excel)",
            any("párrafo" in celda for fila in filas for celda in fila),
        )

        # -- Un PDF sin texto no deja fichero ---------------------------------
        vacio = fitz.open()
        vacio.new_page()
        vacio.save(tmp / "vacio.pdf")
        vacio.close()
        sin_texto = repositorio.abrir(tmp / "vacio.pdf")
        destino = tmp / "no_deberia_existir.xlsx"
        try:
            caso.ejecutar(sin_texto, destino, FormatoTabla.XLSX)
            aviso = ""
        except Exception as error:  # ErrorDominio
            aviso = str(error)
        _check(
            "PDF sin texto: avisa y no deja fichero",
            "escaneado" in aviso and not destino.exists(),
        )

        # En Windows el temporal no se puede borrar con los PDFs aún abiertos.
        registro.cerrar(documento.id)
        registro.cerrar(sin_texto.id)

    print("-" * 68)
    ok = True
    for nombre, cond in _resultados:
        print(f"  [{'OK' if cond else 'FALLO'}] {nombre}")
        ok = ok and cond
    print("-" * 68)
    print("PyMuPDF:", fitz.VersionBind)
    print("RESULTADO:", "OK" if ok else "FALLO")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
