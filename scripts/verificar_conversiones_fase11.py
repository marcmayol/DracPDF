"""Criterio de aceptación funcional de la Fase 11 (más formatos), sin UI ni red.

Salida: PDF → PNG/JPEG/WEBP/TIFF/SVG comprobando la firma de cada fichero, tablas
a CSV y XLSX celda a celda, y ODT y RTF releídos con los lectores de entrada.
Entrada: imágenes, Markdown, HTML, texto, ODT y RTF → PDF, comprobando páginas y
texto extraíble; y un .doc que no se convierte y sí explica qué hacer.

Las conversiones que pasan por Qt necesitan fuentes, así que se fuerza la
plataforma NATIVA (no offscreen).

Uso:
    uv run python scripts/verificar_conversiones_fase11.py
"""

from __future__ import annotations

import csv
import os
import sys
import tempfile
import zipfile
from html import unescape
from pathlib import Path

os.environ.pop("QT_QPA_PLATFORM", None)  # plataforma nativa (con fuentes)
sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # consola cp1252 → utf-8

import fitz  # noqa: E402
from PySide6.QtWidgets import QApplication  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from tests.adapters.generar_fixtures_odt import generar_odt_prueba  # noqa: E402

from lectorpdf.adapters.pymupdf.conversor import ConversorFitz  # noqa: E402
from lectorpdf.adapters.pymupdf.conversor_imagenes import (  # noqa: E402
    ConversorImagenesFitz,
)
from lectorpdf.adapters.pymupdf.document_repository import (  # noqa: E402
    PyMuPDFDocumentRepository,
)
from lectorpdf.adapters.pymupdf.herramientas import PyMuPDFHerramientas  # noqa: E402
from lectorpdf.adapters.pymupdf.registro import RegistroDocumentos  # noqa: E402
from lectorpdf.adapters.pymupdf.tablas import PyMuPDFTablas  # noqa: E402
from lectorpdf.adapters.qt.conversor_texto import ConversorTextoQt  # noqa: E402
from lectorpdf.adapters.qt.odt import odt_a_html  # noqa: E402
from lectorpdf.adapters.qt.rtf import rtf_a_html  # noqa: E402
from lectorpdf.core.domain.conversion import (  # noqa: E402
    A4,
    FORMATOS_SIN_SOPORTE,
    AjusteImagen,
    ConfigPagina,
    FormatoImagen,
)
from lectorpdf.core.domain.tablas import FormatoTabla  # noqa: E402

_resultados: list[tuple[str, bool]] = []

_FIRMAS = {
    FormatoImagen.PNG: b"\x89PNG\r\n\x1a\n",
    FormatoImagen.JPEG: b"\xff\xd8\xff",
    FormatoImagen.WEBP: b"RIFF",
    FormatoImagen.TIFF: b"II*\x00",
}

_FILAS = [["Concepto", "Importe"], ["Alquiler", "750,00"], ["Luz", "120,50"]]


def _check(nombre: str, condicion: bool) -> None:
    _resultados.append((nombre, bool(condicion)))


def _pdf_de_prueba(destino: Path) -> Path:
    """Título grande, párrafo con acentos y una tabla con líneas."""
    doc = fitz.open()
    pagina = doc.new_page(width=595, height=842)
    pagina.insert_text((60, 80), "Informe de prueba", fontsize=24)
    pagina.insert_text((60, 130), "Párrafo con acentos: año, gestión, ñu.", fontsize=11)
    for f, fila in enumerate(_FILAS):
        for c, celda in enumerate(fila):
            rect = fitz.Rect(
                60 + c * 150, 200 + f * 24, 60 + (c + 1) * 150, 224 + f * 24
            )
            pagina.draw_rect(rect, color=(0, 0, 0), width=0.7)
            pagina.insert_text((rect.x0 + 4, rect.y1 - 8), celda, fontsize=10)
    doc.new_page(width=595, height=842).insert_text(
        (60, 80), "Segunda página", fontsize=11
    )
    doc.save(destino)
    doc.close()
    return destino


def _imagen(destino: Path, ancho: int, alto: int, formato: str = "png") -> Path:
    doc = fitz.open()
    pagina = doc.new_page(width=ancho, height=alto)
    pagina.draw_rect(fitz.Rect(0, 0, ancho, alto), color=(0.84, 0.16, 0.16))
    pagina.insert_text((10, alto / 2), f"{ancho}x{alto}", fontsize=10)
    pagina.get_pixmap(dpi=72).save(str(destino), output=formato)
    doc.close()
    return destino


def _texto_pdf(ruta: Path) -> str:
    doc = fitz.open(ruta)
    try:
        return "\n".join(doc[i].get_text("text") for i in range(doc.page_count))
    finally:
        doc.close()


def _paginas(ruta: Path) -> int:
    doc = fitz.open(ruta)
    try:
        return int(doc.page_count)
    finally:
        doc.close()


def main() -> int:
    QApplication.instance() or QApplication([])
    tmp = Path(tempfile.mkdtemp())

    registro = RegistroDocumentos()
    pdf = _pdf_de_prueba(tmp / "origen.pdf")
    doc = PyMuPDFDocumentRepository(registro).abrir(pdf)
    herramientas = PyMuPDFHerramientas(registro)
    conversor = ConversorFitz(registro)
    tablas = PyMuPDFTablas(registro)

    # -- Salida: imágenes -----------------------------------------------------
    for formato in (FormatoImagen.PNG, FormatoImagen.JPEG, FormatoImagen.WEBP):
        rutas = herramientas.exportar_imagenes(
            doc.id, tmp / formato.value, dpi=72, formato=formato
        )
        bien = len(rutas) == 2 and all(
            r.read_bytes().startswith(_FIRMAS[formato]) for r in rutas
        )
        _check(f"PDF→{formato.name}: un fichero por página con su firma", bien)

    tiff = herramientas.exportar_imagenes(
        doc.id, tmp / "tiff", dpi=72, formato=FormatoImagen.TIFF
    )
    from PIL import Image

    with Image.open(tiff[0]) as imagen:
        paginas_tiff = imagen.n_frames
    _check(
        "PDF→TIFF: un solo fichero con todas las páginas",
        len(tiff) == 1 and paginas_tiff == 2,
    )

    svg = herramientas.exportar_imagenes(
        doc.id, tmp / "svg", dpi=72, formato=FormatoImagen.SVG
    )
    contenido_svg = svg[0].read_text(encoding="utf-8")
    _check(
        "PDF→SVG: vectorial y con el texto como texto",
        len(svg) == 2 and "<text" in contenido_svg
        and "Informe de prueba" in unescape(contenido_svg),
    )

    # -- Salida: tablas -------------------------------------------------------
    detectadas = tablas.detectar_tablas(doc.id)
    _check("Tablas: se detecta la tabla con líneas", len(detectadas) == 1)

    csvs = tablas.exportar_tablas(doc.id, tmp / "csv", FormatoTabla.CSV)
    with open(csvs[0], encoding="utf-8-sig", newline="") as fichero:
        filas_csv = list(csv.reader(fichero, delimiter=";"))
    _check("Tablas→CSV: celdas idénticas al documento", filas_csv == _FILAS)

    xlsx = tablas.exportar_tablas(doc.id, tmp / "xlsx", FormatoTabla.XLSX)
    from openpyxl import load_workbook

    libro = load_workbook(xlsx[0])
    filas_xlsx = [[c.value for c in fila] for fila in libro[libro.sheetnames[0]].rows]
    libro.close()
    _check("Tablas→XLSX: celdas idénticas al documento", filas_xlsx == _FILAS)

    vacio = fitz.open()
    vacio.new_page().insert_text((72, 100), "Sin tablas aquí.", fontsize=11)
    ruta_vacio = tmp / "sin_tablas.pdf"
    vacio.save(ruta_vacio)
    vacio.close()
    doc_vacio = PyMuPDFDocumentRepository(registro).abrir(ruta_vacio)
    sin = tablas.exportar_tablas(doc_vacio.id, tmp / "sin", FormatoTabla.CSV)
    _check(
        "Tablas: sin tablas no se escribe ningún fichero",
        sin == [] and not list((tmp / "sin").glob("*.csv")),
    )

    # -- Salida: ODT y RTF ----------------------------------------------------
    odt = tmp / "salida.odt"
    conversor.a_odt(doc.id, odt)
    with zipfile.ZipFile(odt) as paquete:
        mimetype_ok = (
            paquete.namelist()[0] == "mimetype"
            and paquete.getinfo("mimetype").compress_type == zipfile.ZIP_STORED
        )
    html_odt = odt_a_html(odt)
    _check("PDF→ODT: paquete ODF válido", mimetype_ok)
    _check(
        "PDF→ODT: se relee con título, acentos y tabla",
        "Informe de prueba" in html_odt
        and "gestión" in html_odt
        and "750,00" in html_odt,
    )

    rtf = tmp / "salida.rtf"
    conversor.a_rtf(doc.id, rtf)
    crudo = rtf.read_bytes()
    try:
        crudo.decode("ascii")
        ascii_ok = True
    except UnicodeDecodeError:
        ascii_ok = False
    html_rtf = unescape(rtf_a_html(rtf))
    _check("PDF→RTF: ASCII puro (no ASCII escapado)", ascii_ok and crudo.startswith(b"{\\rtf1"))
    _check(
        "PDF→RTF: se relee con título y acentos",
        "Informe de prueba" in html_rtf and "gestión" in html_rtf,
    )

    # -- Entrada: imágenes ----------------------------------------------------
    imagenes = [
        _imagen(tmp / "a.png", 300, 150),
        _imagen(tmp / "b.jpg", 150, 300, formato="jpg"),
    ]
    desde_imgs = tmp / "desde_imagenes.pdf"
    ConversorImagenesFitz().a_pdf(
        imagenes, desde_imgs, AjusteImagen.TAMANO_IMAGEN, ConfigPagina()
    )
    _check("Imágenes→PDF: una página por imagen", _paginas(desde_imgs) == 2)

    desde_imgs_a4 = tmp / "desde_imagenes_a4.pdf"
    ConversorImagenesFitz().a_pdf(
        imagenes, desde_imgs_a4, AjusteImagen.PAGINA_FIJA, ConfigPagina()
    )
    doc_a4 = fitz.open(desde_imgs_a4)
    a4_ok = all(round(p.rect.width) == 595 for p in doc_a4)
    doc_a4.close()
    _check("Imágenes→PDF: página fija A4", a4_ok)

    # -- Entrada: texto, HTML, Markdown, ODT y RTF ----------------------------
    conversor_texto = ConversorTextoQt()

    md = tmp / "notas.md"
    md.write_text("# Título Markdown\n\nPárrafo con **negrita**.\n", encoding="utf-8")
    desde_md = tmp / "desde_md.pdf"
    conversor_texto.a_pdf(md, desde_md, A4)
    texto_md = _texto_pdf(desde_md)
    _check(
        "Markdown→PDF: título y texto sin marcado",
        "Título Markdown" in texto_md and "negrita" in texto_md and "**" not in texto_md,
    )

    html = tmp / "pagina.html"
    html.write_text("<h1>Título HTML</h1><p>Con <em>énfasis</em>.</p>", encoding="utf-8")
    desde_html = tmp / "desde_html.pdf"
    conversor_texto.a_pdf(html, desde_html, A4)
    texto_html = _texto_pdf(desde_html)
    _check(
        "HTML→PDF: se interpreta el marcado",
        "Título HTML" in texto_html and "<p>" not in texto_html,
    )

    txt = tmp / "plano.txt"
    txt.write_bytes("Año de gestión: ñ, á.".encode("cp1252"))  # no UTF-8, a propósito
    desde_txt = tmp / "desde_txt.pdf"
    conversor_texto.a_pdf(txt, desde_txt, A4)
    _check("Texto→PDF: lee cp1252 sin romperse", "Año de gestión" in _texto_pdf(desde_txt))

    odt_entrada = generar_odt_prueba(tmp / "entrada.odt")
    desde_odt = tmp / "desde_odt.pdf"
    conversor_texto.a_pdf(odt_entrada, desde_odt, A4)
    texto_odt = _texto_pdf(desde_odt)
    _check(
        "ODT→PDF: título, negrita, lista y tabla",
        "Informe de prueba" in texto_odt
        and "texto en negrita" in texto_odt
        and "primer punto" in texto_odt
        and "750 EUR" in texto_odt,
    )

    rtf_entrada = tmp / "entrada.rtf"
    rtf_entrada.write_text(
        r"{\rtf1\ansi{\fonttbl{\f0 Calibri;}}Texto \b en negrita\b0  con "
        r"a\'f1o.\par}",
        encoding="latin-1",
    )
    desde_rtf = tmp / "desde_rtf.pdf"
    conversor_texto.a_pdf(rtf_entrada, desde_rtf, A4)
    texto_rtf = _texto_pdf(desde_rtf)
    _check(
        "RTF→PDF: texto y acentos, sin la tabla de fuentes",
        "en negrita" in texto_rtf and "año" in texto_rtf and "Calibri" not in texto_rtf,
    )

    # -- Formatos que no se convierten ---------------------------------------
    aviso = FORMATOS_SIN_SOPORTE.get(".doc")
    _check(
        "Formato sin soporte: el aviso dice qué hacer",
        aviso is not None and "guárdalo como .docx" in aviso,
    )

    print("-" * 68)
    ok = True
    for nombre, cond in _resultados:
        print(f"  [{'OK' if cond else 'FALLO'}] {nombre}")
        ok = ok and cond
    print("-" * 68)
    print("PyMuPDF:", fitz.VersionBind)
    print("RESULTADO:", "OK" if ok else "FALLO")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
